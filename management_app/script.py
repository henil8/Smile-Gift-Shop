import openpyxl
from openpyxl.styles import Font, Alignment

# ==============================
# 1. List of all API names
# ==============================
apis = [
    # Authentication
    "User Sign-Up", "Login", "Logout", "Forgot Password", "Change Password",
    "Verify OTP", "Resend OTP",
    
    # Profile / KYC
    "Get Profile", "Update Profile", "Upload KYC Details", "Update KYC Details",
    "Get Bank Details", "Update Bank Details","Get Address" , "Add Address" ,"Delete Account",

    # Products & Categories
      "Get Categories", "Get Subcategories", "Get Products", "Search Products",
    "Product Details", "Add Product", "Delete Product",'Edit Product',

    "Add-Favourite", "Remove from Favourite", "List-Favourite",

    # Cart
    "Add to Cart", "Update Cart", "Remove from Cart", "Get Cart",

    # Orders
    "Place Order", "Get Order List", "Get Order Details", "Cancel Order",
    "Return Order", "Track Order",

    # Payments
    "Online Payment Offer List",

    # Notifications
    "Get Notifications", "Mark Notification Read",
    
    "Inquiry"
     
]

# ==============================
# 2. Estimated times (dummy values, adjust manually later)
# ==============================
# You can assign times manually or keep blank
default_time = "3-4 hrs"
api_with_time = [(api, default_time) for api in apis]

# ==============================
# 3. Group related APIs
# ==============================
grouped_apis = {
    "Authentication": [a for a in apis if a in ["Register","Login","Logout","Forgot Password","Change Password","Verify OTP","Resend OTP"]],
    "Profile & KYC": [a for a in apis if "Profile" in a or "KYC" in a or "Bank" in a or "address" in a or "Account" in a],
    "Products": [a for a in apis if "Product" in a  or "Wishlist" in a],
    "Categories": [a for a in apis if "Categories" in a  or "Wishlist" in a],
    "Favourites": [a for a in apis if "Favourite" in a],
    "Cart": [a for a in apis if "Cart" in a],
    "Orders": [a for a in apis if "Order" in a or "Track" in a],
    "Payments": [a for a in apis if "Payment" in a],
    "Notifications": [a for a in apis if "Notification" in a],
    "Support": [a for a in apis if "Inquiry" in a or "Ticket" in a],
}

# ==============================
# 4. Create Excel
# ==============================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "APIs"

# Heading
ws.merge_cells("A1:B1")
ws["A1"] = "Smile Gift Shop Mobile APIs"
ws["A1"].font = Font(size=14, bold=True)
ws["A1"].alignment = Alignment(horizontal="center")

row = 3
for group, api_list in grouped_apis.items():
    # Write group name
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value=group).font = Font(bold=True, underline="single")
    row += 1
    
    # Write APIs under the group
    for api in api_list:
        ws.cell(row=row, column=1, value=api)
        ws.cell(row=row, column=2, value=default_time)  # default time
        row += 1
    row += 1  # add spacing after each group

# Save file
wb.save("Smile_Gift_Shop_Mobile_APIs.xlsx")
print("Excel file created: Smile_Gift_Shop_Mobile_APIs.xlsx")
